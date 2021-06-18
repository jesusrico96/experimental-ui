import os.path
from tkinter import *
from openpyxl import *
from PIL import *
import statistics
import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd
from tkinter import filedialog

global M
global SD
global median
global MAD
M = 0
SD = 0
median = 0
MAD = 0
dirname = os.path.dirname(__file__)


def browsefiles():
    global filename
    filename = filedialog.askopenfilename(initialdir=sys.path[0],
                                          title="Escolle un arquivo",
                                          filetypes=(("Ficheiro de Excel",
                                                      "*.xlsx"),
                                                     ("Todos os formatos",
                                                      "*.*")))

    # Etiqueta do ficheiro utilizado
    label_file_explorer.configure(text=filename)
    print(filename)


def guillotina():
    # Abrir o arquivo cos datos e un workbook paralelo de destino dos mesmos
    my_file = load_workbook(filename=filename)
    try:
        my_sheet = my_file.worksheets[1]
    except:
        my_sheet = my_file.worksheets[0]
    global wb
    global datos
    wb = Workbook()
    datos = wb.active

    # Contar filas e columnas do arquivo orixinal e pegalas no arquivo de destino
    mr = my_sheet.max_row
    mc = my_sheet.max_column

    for i in range(1, mr + 1):
        for j in range(1, mc + 1):
            c = my_sheet.cell(row=i, column=j)
            datos.cell(row=i, column=j).value = c.value
    print("Raw target file created!")

    # Eliminar os datos innecesarios
    datos.delete_cols(4)
    datos.delete_cols(7, 3)
    datos.delete_cols(16, 7)
    datos.delete_cols(18)
    datos.delete_cols(20)
    datos.delete_cols(30, 4)
    datos.delete_cols(34)
    datos.delete_cols(37, 2)
    datos.delete_cols(39)
    datos.delete_cols(40, 9)

    # Identificador de fin de ensaio e guillotina de filas sobrantes
    vali = 0
    for i in range(2, mr + 1):
        fin = datos.cell(i, 39).value
        if fin == 1:
            vali = i
    global mr2
    mr2 = vali
    datos.delete_rows(mr2 + 1, mr + 1)
    print("End of Test located and data trimmed.")

    # Conversor de CO a CO (ppm) e vacía cela en caso de CO<200
    datos.cell(1, 40).value = "CO (ppm) corr"
    for i in range(2, mr2 + 1):
        ttl = datos.cell(i, 31).value * 10000
        if ttl > 200:
            datos.cell(i, 40).value = ttl
    print("Corrected CO (ppm) and limited to over 200.")

    # Cálculo de TL promedios
    datos.cell(1, 41).value = "TL Alta (1-3)"
    for i in range(2, mr2 + 1):
        tlalta = (datos.cell(i, 20).value + datos.cell(i, 21).value + datos.cell(i, 22).value) / 3
        if tlalta > 0:
            datos.cell(i, 41).value = tlalta

    datos.cell(1, 42).value = "TL Media (4-6)"
    for i in range(2, mr2 + 1):
        tlmedia = (datos.cell(i, 23).value + datos.cell(i, 24).value + datos.cell(i, 25).value) / 3
        if tlmedia > 0:
            datos.cell(i, 42).value = tlmedia

    datos.cell(1, 43).value = "TL Baixa (7-9)"
    for i in range(2, mr2 + 1):
        tlbaixa = (datos.cell(i, 26).value + datos.cell(i, 27).value + datos.cell(i, 28).value) / 3
        if tlbaixa > 0:
            datos.cell(i, 43).value = tlbaixa

    print("Mean temperatures in Combustion Chamber calculated.")

    # Eliminación de datos sobrantes
    datos.delete_cols(20, 9)  # TL1 a TL9
    datos.delete_cols(22)  # CO

    # Pechar o arquivo orixinal e gardar un Excel cos datos de traballo
    my_file.close()
    wb.save(filename="datos.xlsx")  # Abre o arquivo de datos e garda unha copia da folla de interese.

    print("Guillotinado!")  # Amosa unha confirmación da execución en consola.


# Fixación tempos
def time():
    global Ti
    global Tf
    try:
        Ti = float(Tempiniw.get())
        Tf = float(Tempfinw.get())
        print("Time set to interval between " + str(Ti) + " and " + str(Tf) + " s.")

    except:
        Ti = "tiempo1"
        Tf = "tiempo2"
        print("Time restrictions removed.")


# Media e desv. típica no rango de tempos
def MDT():
    try:
        # Tempo:
        tempo = list(datos.columns)[1]  # Esto é unha tupla que contén cellObj
        # Lista que contén todos os tempos en orde:
        tempo2 = []
        # Lista que contén todos os tempos solicitados:
        tempo3 = []
        # Lista que contén todos os índices dos tempos solicitados na lista global,
        # de xeito que se podan relacionar con outros parámetros:
        tindex = []
        for cellObj in tempo:
            x = cellObj.value
            tempo2.append(x)
        tempo2.pop(0)
        for x in tempo2:
            if Tf > x > Ti:
                tempo3.append(x)
                tindex.append(tempo2.index(x))
            elif x > Tf:
                break

        # Variable a analizar
        analito = list(datos.columns)[opcionso1.index(variableo1.get())]
        analito2 = []
        analito3 = []
        for cellObj in analito:
            x = cellObj.value
            analito2.append(x)
        analito2.pop(0)

        for x in analito2:
            for n in tindex:
                if analito2.index(x) == n and len(analito3) < len(tindex):
                    analito3.append(x)
                elif len(analito3) == len(tindex):
                    break
                elif x is None:
                    break

        M = round(statistics.mean(analito3), 2)
        median = round(statistics.median(analito3), 2)
        SD = round(statistics.stdev(analito3), 2)
        series = pd.Series(analito3)
        MAD = round(series.mad())

    except:
        # Tempo:
        tempo = list(datos.columns)[1]  # Esto é unha tupla que contén cellObj
        tempo2 = []  # Esta lista conterá todos os tempos en orde
        for cellObj in tempo:
            x = cellObj.value
            tempo2.append(x)
        tempo2.pop(0)

        # Variable a analizar
        analito = list(datos.columns)[opcionso1.index(variableo1.get())]
        analito2 = []
        for cellObj in analito:
            x = cellObj.value
            analito2.append(x)
        analito2.pop(0)

        M = round(statistics.mean(analito2), 2)
        median = round(statistics.median(analito2), 2)
        SD = round(statistics.stdev(analito2), 2)
        series = pd.Series(analito2)
        MAD = round(series.mad())

    label_Mean.configure(text="Mean" + " " + str(M))
    label_Median.configure(text="Median:" + " " + str(median))
    label_SD.configure(text="Desv. Est." + " " + str(SD))
    label_MAD.configure(text="Mean Abs. Dev.:" + " " + str(MAD))


# Gráficos
def grafico():
    # Se a checkbox está desmarcada, fai gráficos só cunha variable en Y. Dentro de cada tipo de gráfico
    # hai un bucle para comprobar se se pide o gráfico nun tempo limitado ou para todo o set de datos.
    if bo2.get() == 0:
        try:
            # Tempo:
            # Esto é unha tupla que contén cellObj:
            abscisa = list(datos.columns)[opcionsa1.index(variablea1.get())]
            # Esta lista conterá todos os tempos en orde:
            abscisa2 = []
            # Esta lista conterá só os tempos solicitados:
            abscisa3 = []
            # Lista que contén todos os índices dos tempos solicitados na lista global,
            # de xeito que se podan relacionar con outros parámetros:
            tindex = []
            for cellObj in abscisa:
                x = cellObj.value
                abscisa2.append(x)
            abscisa2.pop(0)
            for x in abscisa2:
                if Tf > x > Ti:
                    abscisa3.append(x)
                    tindex.append(abscisa2.index(x))
                elif x > Tf:
                    break

            # Variable a analizar
            analito = list(datos.columns)[opcionso1.index(variableo1.get())]
            analito2 = []
            analito3 = []
            for cellObj in analito:
                x = cellObj.value
                analito2.append(x)
            analito2.pop(0)

            for x in analito2:
                for n in tindex:
                    if analito2.index(x) == n and len(analito3) < len(tindex):
                        analito3.append(x)
                    elif len(analito3) == len(tindex):
                        break

            # plot
            plt.plot(abscisa3, analito3, label=variableo1.get(), linewidth=0,
                     marker='.', markerfacecolor='blue', markersize=3)

            # naming the x axis
            plt.xlabel(variablea1.get())
            # naming the y axis
            plt.ylabel(variableo1.get())
            # Legend
            plt.legend()

            # giving a title to my graph
            plt.title(str(variablea1.get()) + " vs. " + str(variableo1.get()))

        except:
            # Tempo:
            abscisa = list(datos.columns)[opcionsa1.index(variablea1.get())]  # Esto é unha tupla que contén cellObj
            abscisa2 = []  # Esta lista conterá todos os tempos en orde
            for cellObj in abscisa:
                x = cellObj.value
                abscisa2.append(x)
            abscisa2.pop(0)

            # Variable a analizar
            analito = list(datos.columns)[opcionso1.index(variableo1.get())]
            analito2 = []
            for cellObj in analito:
                x = cellObj.value
                analito2.append(x)
            analito2.pop(0)

            # plot
            plt.plot(abscisa2, analito2, label=variableo1.get(), linewidth=0,
                     marker='.', markerfacecolor='blue', markersize=3)

            plt.xlabel(variablea1.get())
            plt.ylabel(variableo1.get())
            plt.legend()

            plt.title(str(variablea1.get()) + " vs. " + str(variableo1.get()))
    else:
        try:
            # Tempo:
            abscisa = list(datos.columns)[opcionsa1.index(variablea1.get())]
            abscisa2 = []
            abscisa3 = []
            tindex = []
            for cellObj in abscisa:
                x = cellObj.value
                abscisa2.append(x)
            abscisa2.pop(0)
            for x in abscisa2:
                if Tf > x > Ti:
                    abscisa3.append(x)
                    tindex.append(abscisa2.index(x))
                elif x > Tf:
                    break

            # Variables a analizar
            analitoa = list(datos.columns)[opcionso1.index(variableo1.get())]
            analitoa2 = []
            analitoa3 = []
            for cellObj in analitoa:
                x = cellObj.value
                analitoa2.append(x)
            analitoa2.pop(0)

            for x in analitoa2:
                for n in tindex:
                    if analitoa2.index(x) == n and len(analitoa3) < len(tindex):
                        analitoa3.append(x)
                    elif len(analitoa3) == len(tindex):
                        break

            analitob = list(datos.columns)[opcionso2.index(variableo2.get())]
            analitob2 = []
            analitob3 = []
            for cellObj in analitob:
                x = cellObj.value
                analitob2.append(x)
            analitob2.pop(0)

            for x in analitob2:
                for n in tindex:
                    if analitob2.index(x) == n and len(analitob3) < len(tindex):
                        analitob3.append(x)
                    elif len(analitob3) == len(tindex):
                        break

            # plot
            plt.plot(abscisa3, analitoa3, label=variableo1.get(), linewidth=0,
                     marker='.', markerfacecolor='blue', markersize=3)

            plt.plot(abscisa3, analitob3, label=variableo2.get(), linewidth=0,
                     marker='.', markerfacecolor='r', markersize=3)

            plt.xlabel(variablea1.get())
            plt.ylabel(variableo1.get() + " and " + variableo2.get())
            plt.legend()

            plt.title(str(variablea1.get()) + " vs. " + str(variableo1.get()) + " and " + variableo2.get())

        except:
            # Tempo:
            abscisa = list(datos.columns)[opcionsa1.index(variablea1.get())]  # Esto é unha tupla que contén cellObj
            abscisa2 = []  # Esta lista conterá todos os tempos en orde
            for cellObj in abscisa:
                x = cellObj.value
                abscisa2.append(x)
            abscisa2.pop(0)

            # Variables a analizar
            analitoa = list(datos.columns)[opcionso1.index(variableo1.get())]
            analitoa2 = []
            for cellObj in analitoa:
                x = cellObj.value
                analitoa2.append(x)
            analitoa2.pop(0)

            analitob = list(datos.columns)[opcionso2.index(variableo2.get())]
            analitob2 = []
            for cellObj in analitob:
                x = cellObj.value
                analitob2.append(x)
            analitob2.pop(0)

            # plot
            plt.plot(abscisa2, analitoa2, label=variableo1.get(), linewidth=0,
                     marker='.', markerfacecolor='blue', markersize=3)

            plt.plot(abscisa2, analitob2, label=variableo2.get(), linewidth=0,
                     marker='.', markerfacecolor='r', markersize=3)

            plt.xlabel(variablea1.get())
            plt.ylabel(variableo1.get() + " and " + variableo2.get())
            plt.legend()

            plt.title(str(variablea1.get()) + " vs. " + str(variableo1.get()) + " and " + variableo2.get())

    # function to show the plot
    plt.show()


# Boxplots
def boxplot():
    if bv3.get() == 0:
        print("1 independent variable detected, drawing 1 Boxplot.")
        try:
            # Tempo:
            abscisa = list(datos.columns)[opcionsv1.index(variable1.get())]
            abscisa2 = []
            abscisa3 = []
            tindex = []
            for cellObj in abscisa:
                x = cellObj.value
                abscisa2.append(x)
            abscisa2.pop(0)

            print("Applying temporal restraint to X Axis")
            for x in abscisa2:
                if Tf > x > Ti:
                    abscisa3.append(x)
                    tindex.append(abscisa2.index(x))
                elif x > Tf:
                    break
            print("Temporal limitation to X Axis successful")

            # Variable a analizar
            analito = list(datos.columns)[opcionsv2.index(variable2.get())]
            analito2 = []
            analito3 = []
            for cellObj in analito:
                x = cellObj.value
                analito2.append(x)
            analito2.pop(0)

            print("Applying temporal restraint to Y Axis")
            for x in analito2:
                for n in tindex:
                    if analito2.index(x) == n and len(analito3) < len(tindex):
                        analito3.append(x)
                    elif len(analito3) == len(tindex):
                        break
            print("Temporal limitation to Y Axis successful")

            sns.boxplot(data=analito3)

            # naming the axis
            plt.xlabel(variable2.get())

        except:
            print("No temporal restriction detected.")

            # Variable a analizar
            analito = list(datos.columns)[opcionsv2.index(variable2.get())]
            analito2 = []
            for cellObj in analito:
                x = cellObj.value
                analito2.append(x)
            analito2.pop(0)
            print("Variable 2 data extracted.")

            sns.boxplot(data=analito2)

            # naming the axes

            plt.xlabel(variable2.get())
    else:
        print("2 independent variables detected, drawing 2 Boxplots.")
        try:
            # Tempo:
            abscisa = list(datos.columns)[opcionsv1.index(variable1.get())]
            abscisa2 = []
            abscisa3 = []
            tindex = []
            for cellObj in abscisa:
                x = cellObj.value
                abscisa2.append(x)
            abscisa2.pop(0)
            print("Applying temporal restraint to X Axis")
            for x in abscisa2:
                if Tf > x > Ti:
                    abscisa3.append(x)
                    tindex.append(abscisa2.index(x))
                elif x > Tf:
                    break
            print("Temporal limitation to X Axis successful")

            # Variables a analizar
            analitoa = list(datos.columns)[opcionsv2.index(variable2.get())]
            analitoa2 = []
            analitoa3 = []
            for cellObj in analitoa:
                x = cellObj.value
                analitoa2.append(x)
            analitoa2.pop(0)

            print("Applying temporal limitation to Y1 Axis.")
            for x in analitoa2:
                for n in tindex:
                    if analitoa2.index(x) == n and len(analitoa3) < len(tindex):
                        analitoa3.append(x)
                    elif len(analitoa3) == len(tindex):
                        break
            print("Temporal limitation to Y1 Axis successful")

            analitob = list(datos.columns)[opcionsv3.index(variable3.get())]
            analitob2 = []
            analitob3 = []
            for cellObj in analitob:
                x = cellObj.value
                analitob2.append(x)
            analitob2.pop(0)

            print("Applying temporal limitation to Y2 Axis.")
            for x in analitob2:
                for n in tindex:
                    if analitob2.index(x) == n and len(analitob3) < len(tindex):
                        analitob3.append(x)
                    elif len(analitob3) == len(tindex):
                        break
            print("Temporal limitation to Y2 Axis successful")

            df = pd.DataFrame(list(zip(analitoa3, analitob3)))
            df.columns = [str(variable2.get()), str(variable3.get())]

            sns.boxplot(data=df)

            # naming the y axis
            plt.ylabel(variable2.get() + " and " + variable3.get())

        except:
            print("No temporal restriction detected.")

            # Variables a analizar
            analitoa = list(datos.columns)[opcionsv2.index(variable2.get())]
            analitoa2 = []
            for cellObj in analitoa:
                x = cellObj.value
                analitoa2.append(x)
            analitoa2.pop(0)
            print("Variable 2 data acquired.")

            analitob = list(datos.columns)[opcionsv3.index(variable3.get())]
            analitob2 = []
            for cellObj in analitob:
                x = cellObj.value
                analitob2.append(x)
            analitob2.pop(0)
            print("Variable 3 data obtained.")

            df = pd.DataFrame(list(zip(analitoa2, analitob2)))
            df.columns = [str(variable2.get()), str(variable3.get())]

            sns.boxplot(data=df)

            # naming the y axis
            plt.ylabel(variable2.get() + " and " + variable3.get())
    # function to show the plot
    plt.show()


# Ventana principal
window = Tk()

# Titulo da ventá principal
window.title('FTDE: BICO')

# Tamaño da ventá principal
window.geometry("725x400")

# Cor de fondo da venta principal
window.config(background="white")

# Imos crear algúns frames de fondo
top_frame = Frame(window, bg="black", width=450, height=50, pady=3)
top_frame2 = Frame(window, bg="alice blue", width=450, height=50, pady=6)
middle_frame = Frame(window, bg="black", width=450, height=400, pady=3)
low_frame = Frame(window, bg="gray84", width=450, height=50, pady=3)

# Disposición dos frames de fondo
window.grid_rowconfigure(1, weight=1)
window.grid_columnconfigure(0, weight=1)

top_frame.grid(row=0, sticky="ew")
top_frame2.grid(row=1, sticky="ew")
middle_frame.grid(row=2, sticky="nsew")
low_frame.grid(row=3, sticky="ew")

# Creación e disposición dos frames de fronte (centro)
middle_frame.grid_rowconfigure(0, weight=1)
middle_frame.grid_columnconfigure(1, weight=1)

framea = Frame(middle_frame, bg='alice blue', width=250, height=300, padx=3, pady=3)
frameb = Frame(middle_frame, bg='alice blue', width=250, height=300, padx=3, pady=3)

framea.grid(row=0, column=0, columnspan=2, sticky="nsew", padx=3)
frameb.grid(row=0, column=2, sticky="ns", padx=3)

# Etiquetas do Explorador de Arquivos
label_file_explorer = Label(top_frame,
                            text="Ferramenta de tratamento de datos experimentais do GTE: versión BICO",
                            width=110, height=4,
                            fg="blue")

# Etiqueta co logo do GTE
photogte = PhotoImage(file=os.path.join(dirname, "imaxes", "gte2.png"))

photoimagegte = photogte.subsample(3, 3)

label_gte = Label(top_frame2,
                  image=photoimagegte)

# Botóns do Explorador de Arquivos
button_explore = Button(top_frame2,
                        text="Navegar arquivos",
                        command=browsefiles)

button_exit = Button(low_frame,
                     text="Saír",
                     fg="red",
                     command=exit)

# Etiqueta autoría
label_auth = Label(low_frame,
                   text="GNU-GPL   2021   Juan Jesús Rico Fuentes",
                   fg="black")

# Botón de guillotina
photo = PhotoImage(
    file=os.path.join(dirname, "imaxes", "guillo.png"))

photoimage = photo.subsample(17, 17)

button_guillotina = Button(top_frame2,
                           text="Guillotina",
                           image=photoimage,
                           command=guillotina)

# Widget entrada de tempos:

# Etiqueta para tempo inicial
Tempini_label = Label(top_frame2,
                      text='Tempo inicial')

# Entrada do tempo incial
Tempiniw = Entry(top_frame2,
                 width=10)

# Etiqueta para tempo final
Tempfin_label = Label(top_frame2,
                      text='Tempo final')

# Entrada do tempo final
Tempfinw = Entry(top_frame2,
                 width=10)

# Fixar tempos
button_tempos = Button(top_frame2,
                       text="Fixar tempos",
                       command=time)

# Etiqueta gráficos
label_partial_graph = Label(framea,
                            width=70, height=1,
                            text="Gráficos")

# Media e desv. típ.
button_MDT = Button(framea,
                    text="MDT (y)",
                    command=MDT)

# Etiqueta coa media
label_Mean = Label(framea,
                   text="Media:" + " " + str(M),
                   fg="blue")

# Etiqueta coa mediana
label_Median = Label(framea,
                     text="Median:" + " " + str(median),
                     fg="blue")

# Etiqueta coa standard dev.
label_SD = Label(framea,
                 text="Std. Dev.:" + " " + str(SD),
                 fg="blue")

# Etiqueta coa MAD
label_MAD = Label(framea,
                  text="Mean Abs. Dev.:" + " " + str(MAD),
                  fg="blue")

# Botón de Gráficos
button_graph = Button(framea,
                      text="Gráfico",
                      width=10, height=2,
                      command=grafico)

# Botón de Boxplots
button_boxp1 = Button(frameb,
                      text="Box Plot 1",
                      width=10, height=2,
                      command=boxplot)

# Etiqueta de boxplots
label_global_graph = Label(frameb,
                           width=20, height=1,
                           text="Boxplots")

# Etiqueta "Eixo X" grafico
label_a1 = Label(framea,
                 text="Eixo X:")

# Etiqueta "Eixo Y" grafico
label_o1 = Label(framea,
                 text="Eixo Y:")

# Etiqueta "Eixo Y 2" grafico e checkbutton para activala
label_o2 = Label(framea,
                 text="Eixo Y 2:")

bo2 = IntVar()
button_o2 = Checkbutton(framea, text='', variable=bo2, onvalue=1, offvalue=0)

# Etiqueta "Var 1" boxplot
label_v1 = Label(frameb,
                 text="Variable 1 (t):")

# Etiqueta "Var 2" boxplot
label_v2 = Label(frameb,
                 text="Variable 2:")

# Etiqueta "Var 3" boxplot con botón para activar
label_v3 = Label(frameb,
                 text="Variable 3:")
bv3 = IntVar()
button_v3 = Checkbutton(frameb, text='', variable=bv3, onvalue=1, offvalue=0)

# Widget para selección de variable a plottear. A orde das variables é a do LabView do Drop, así que haberá que
# axustala se se cambia de instalación

# Lista e widget de selección para abscisas
opcionsa1 = ["Hora", "Tempo (s)", "Masa pellet", "Tciclo", "Ton", "Toff", "QG1", "QG2", "SumQG", "TFGR1", "TFGR2",
             "TH1", "TH2", "TH3", "TC", "TACi", "TACo", "TALi", "TALo", "O2", "O2ref", "CO2corr", "NOcorr", "Lambda",
             "Lambda Postc", "O2 Postc", "FGR", "Qrec", "Alim. pellet", "CO (ppm) corr",
             "TL Alta (1-3)", "TL Media (4-6)", "TL Baixa (7-9)"]

variablea1 = StringVar(window)
variablea1.set(opcionsa1[1])

menua1 = OptionMenu(framea, variablea1, *opcionsa1)

# Lista e widget de selección para ordenadas 1
opcionso1 = ["Hora", "Tempo (s)", "Masa pellet", "Tciclo", "Ton", "Toff", "QG1", "QG2", "SumQG", "TFGR1", "TFGR2",
             "TH1", "TH2", "TH3", "TC", "TACi", "TACo", "TALi", "TALo", "O2", "O2ref", "CO2corr", "NOcorr", "Lambda",
             "Lambda Postc", "O2 Postc", "FGR", "Qrec", "Alim. pellet", "CO (ppm) corr",
             "TL Alta (1-3)", "TL Media (4-6)", "TL Baixa (7-9)"]

variableo1 = StringVar(window)
variableo1.set(opcionso1[2])

menuo1 = OptionMenu(framea, variableo1, *opcionso1)

# Lista e widget de selección para ordenadas 2
opcionso2 = ["Hora", "Tempo (s)", "Masa pellet", "Tciclo", "Ton", "Toff", "QG1", "QG2", "SumQG", "TFGR1", "TFGR2",
             "TH1", "TH2", "TH3", "TC", "TACi", "TACo", "TALi", "TALo", "O2", "O2ref", "CO2corr", "NOcorr", "Lambda",
             "Lambda Postc", "O2 Postc", "FGR", "Qrec", "Alim. pellet", "CO (ppm) corr",
             "TL Alta (1-3)", "TL Media (4-6)", "TL Baixa (7-9)"]

variableo2 = StringVar(window)
variableo2.set(opcionso2[3])

menuo2 = OptionMenu(framea, variableo2, *opcionso2)

# Lista e widget de selección para Variable 1 Boxplot
opcionsv1 = ["Hora", "Tempo (s)", "Masa pellet", "Tciclo", "Ton", "Toff", "QG1", "QG2", "SumQG", "TFGR1", "TFGR2",
             "TH1", "TH2", "TH3", "TC", "TACi", "TACo", "TALi", "TALo", "O2", "O2ref", "CO2corr", "NOcorr", "Lambda",
             "Lambda Postc", "O2 Postc", "FGR", "Qrec", "Alim. pellet", "CO (ppm) corr",
             "TL Alta (1-3)", "TL Media (4-6)", "TL Baixa (7-9)"]

variable1 = StringVar(window)
variable1.set(opcionsv1[1])

menuv1 = OptionMenu(frameb, variable1, *opcionsv1)

# Lista e widget de selección para Variable 2 Boxplot
opcionsv2 = ["Hora", "Tempo (s)", "Masa pellet", "Tciclo", "Ton", "Toff", "QG1", "QG2", "SumQG", "TFGR1", "TFGR2",
             "TH1", "TH2", "TH3", "TC", "TACi", "TACo", "TALi", "TALo", "O2", "O2ref", "CO2corr", "NOcorr", "Lambda",
             "Lambda Postc", "O2 Postc", "FGR", "Qrec", "Alim. pellet", "CO (ppm) corr",
             "TL Alta (1-3)", "TL Media (4-6)", "TL Baixa (7-9)"]

variable2 = StringVar(window)
variable2.set(opcionsv2[2])

menuv2 = OptionMenu(frameb, variable2, *opcionsv2)

# Lista e widget de selección para Variable 3 Boxplot
opcionsv3 = ["Hora", "Tempo (s)", "Masa pellet", "Tciclo", "Ton", "Toff", "QG1", "QG2", "SumQG", "TFGR1", "TFGR2",
             "TH1", "TH2", "TH3", "TC", "TACi", "TACo", "TALi", "TALo", "O2", "O2ref", "CO2corr", "NOcorr", "Lambda",
             "Lambda Postc", "O2 Postc", "FGR", "Qrec", "Alim. pellet", "CO (ppm) corr",
             "TL Alta (1-3)", "TL Media (4-6)", "TL Baixa (7-9)"]

variable3 = StringVar(window)
variable3.set(opcionsv3[3])

menuv3 = OptionMenu(frameb, variable3, *opcionsv3)

# Grid method is chosen for placing
# the widgets at respective positions
# in a table like structure by
# specifying rows and columns

label_file_explorer.grid(column=1, row=1)

button_explore.grid(column=1, row=1, rowspan=2, pady=10, padx=5)
button_guillotina.grid(column=2, row=1, rowspan=2, pady=10, padx=5)

label_gte.grid(column=8, row=1, rowspan=2, padx=5)

top_frame2.grid_columnconfigure(7, weight=4)

button_exit.grid(column=3, row=7, padx=5)
low_frame.grid_columnconfigure(4, weight=4)
label_auth.grid(column=5, row=7, padx=5, sticky=E)

# Widget entrada de tempos


Tempini_label.grid(column=3, row=1, pady=5, padx=5)
Tempiniw.grid(column=4, row=1, pady=5, padx=5)
Tempfin_label.grid(column=3, row=2, pady=5, padx=5)
Tempfinw.grid(column=4, row=2, pady=5, padx=5)
button_tempos.grid(column=5, row=1, rowspan=2, pady=5, padx=5)

# Gráficos
label_partial_graph.grid(column=1, row=1, columnspan=6, pady=5, padx=5)

button_MDT.grid(column=4, row=2, rowspan=4, pady=5, padx=5)
label_Mean.grid(column=5, row=2, pady=5, padx=5)
label_Median.grid(column=5, row=3, pady=5, padx=5)
label_SD.grid(column=5, row=4, pady=5, padx=5)
label_MAD.grid(column=5, row=5, pady=5, padx=5)

# Widget de lista desplegable abscisas

label_a1.grid(column=2, row=2, pady=5, padx=5, sticky=E)
menua1.grid(column=3, row=2, pady=5, padx=5)

# Widget de lista desplegable ordenadas 1

label_o1.grid(column=2, row=3, pady=5, padx=5, sticky=E)
menuo1.grid(column=3, row=3, pady=5, padx=5)

# Widget de lista desplegable ordenadas 2

label_o2.grid(column=2, row=4, pady=5, padx=5, sticky=E)
menuo2.grid(column=3, row=4, pady=5, padx=5)
button_o2.grid(column=2, row=4, pady=5, sticky=W)

button_graph.grid(column=1, row=5, columnspan=3, pady=5)

# Widget de Variable temporal Boxplot

label_v1.grid(column=2, row=1, pady=5, padx=5, sticky=E)
label_global_graph.grid(column=2, row=0, columnspan=2, pady=5, padx=5, sticky="N")
menuv1.grid(column=3, row=1, pady=8, padx=5)
frameb.rowconfigure((1, 5), weight=1)

# Widget de Variable 2 Boxplot

label_v2.grid(column=2, row=3, pady=5, padx=5, sticky=E)
menuv2.grid(column=3, row=3, pady=8, padx=5)

# Widget de Variable 3 Boxplot + Checkbutton

label_v3.grid(column=2, row=4, pady=5, padx=5, sticky=E)
button_v3.grid(column=2, row=4, pady=5, sticky=W)
menuv3.grid(column=3, row=4, pady=5, padx=5)

button_boxp1.grid(column=1, row=5, columnspan=3, pady=5, padx=5)

# Let the window wait for any events
window.mainloop()
