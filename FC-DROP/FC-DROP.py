import sys
import os.path
import openpyxl as opx
from tkinter import *
import pandas as pd
# Explorador de arquivos
from tkinter import filedialog


def converter():
    # Na carpeta do script, lista os arquivos que contén:
    cwd = os.path.abspath('')
    files = os.listdir(cwd)

    for file in files:
        if file.endswith('.lvm'):
            my_file = open(file)
            string_list = my_file.readlines()
            string_list[0:21] = ""  # Crea 0 liñas baleiras no espazo ocupado antes por 21 liñas.
            my_file.close()

            my_file = open(file, "w")
            new_file_contents = "".join(string_list)
            my_file.write(new_file_contents)
            my_file.close()

            my_file = open(file)
            df = pd.read_csv(my_file, sep="\t", decimal=',', header=0)
            df.to_excel(file+".xlsx")
            my_file.close()

            my_file = opx.load_workbook(filename=file+".xlsx")
            my_sheet = my_file.worksheets[0]
            my_sheet.delete_cols(28, 8)
            my_sheet.delete_cols(20)
            my_sheet.delete_cols(12, 3)
            my_sheet.delete_cols(1, 2)

            # Contar filas e columnas do arquivo
            mr = my_sheet.max_row
            mc = my_sheet.max_column

            # Corrección de gases
            my_sheet.cell(1, 23).value = "NO (ppm) corr"
            for i in range(2, mr + 1):
                ttl = my_sheet.cell(i, 10).value * ((21-10)/(21-my_sheet.cell(i, 12).value))
                my_sheet.cell(i, 23).value = ttl
                # my_sheet.move_range("W1:W"+str(mr), cols=-13)
            print("NO corrected!")

            my_sheet.cell(1, 24).value = "NOx (ppm) corr"
            for i in range(2, mr + 1):
                ttl = my_sheet.cell(i, 11).value * ((21 - 10) / (21 - my_sheet.cell(i, 12).value))
                my_sheet.cell(i, 24).value = ttl
                # my_sheet.move_range("X:X", cols=-13)
            print("NOx corrected!")

            my_sheet.cell(1, 25).value = "CO (ppm) corr"
            for i in range(2, mr + 1):
                ttl = my_sheet.cell(i, 13).value * ((21 - 10) / (21 - my_sheet.cell(i, 12).value))*10000
                my_sheet.cell(i, 25).value = ttl
                # my_sheet.move_range("Y:Y", cols=-12)
            print("CO corrected!")

            my_sheet.cell(1, 26).value = "CO2 (%) corr"
            for i in range(2, mr + 1):
                ttl = my_sheet.cell(i, 14).value * ((21 - 10) / (21 - my_sheet.cell(i, 12).value))
                my_sheet.cell(i, 26).value = ttl
                # my_sheet.move_range("Z:Z", cols=-12)
            print("CO2 corrected!")

            my_sheet.delete_cols(13, 2)
            my_sheet.delete_cols(10, 2)

            my_file.save(filename=file+".xlsx")
    print("Files converted!")


# Ventana principal
window = Tk()

# Titulo da ventá principal
window.title('File Converter: DROP')

# Tamaño da ventá principal
window.geometry("550x220")

# Cor de fondo da venta principal
window.config(background="white")

# Imos crear algúns frames de fondo
top_frame = Frame(window, bg="black", width=450, height=50, pady=3)
top_frame2 = Frame(window, bg="alice blue", width=450, height=50, pady=6)
low_frame = Frame(window, bg="gray84", width=450, height=50, pady=3)

# Disposición dos frames de fondo
window.grid_rowconfigure(1, weight=1)
window.grid_columnconfigure(0, weight=1)

top_frame.grid(row=0, sticky="ew")
top_frame2.grid(row=1, sticky="ew")
low_frame.grid(row=3, sticky="ew")

# Etiquetas do Explorador de Arquivos
label_file_explorer = Label(top_frame,
                            text="Ferramenta conversión de datos .lvm a .xlsx: versión DROP",
                            width=80, height=4,
                            fg="blue")

# Etiqueta co logo do GTE
dirname = os.path.dirname(__file__)
photogte = PhotoImage(file=os.path.join(dirname, "imaxes", "gte2.png"))

photoimagegte = photogte.subsample(3, 3)

label_gte = Label(top_frame2,
                  image=photoimagegte,)

# Botóns do Explorador de Arquivos

button_exit = Button(low_frame,
                     text="Saír",
                     fg="red",
                     command=exit)

# Botón de conversión
photo = PhotoImage(file=os.path.join(dirname, "imaxes", "conv.png"))

photoimage = photo.subsample(25, 25)

button_converter = Button(top_frame2,
                             text="Converter",
                             image=photoimage,
                             command=converter)

# Etiqueta Descrición
label_desc = Label(top_frame2,
                   text="Esta ferramenta converterá todos os ficheiros .lvm \n da carpeta onde se atope "
                        "en ficheiros .xlsx e, ademais, \n recortará os datos e corrixirá os gases ó 10%",
                   fg="black")

# Etiqueta autoría
label_auth = Label(low_frame,
                   text="GNU-GPL   2021   Juan Jesús Rico Fuentes",
                   fg="black")

# Grid

label_file_explorer.grid(column=2, row=1, columnspan=3)

label_desc.grid(column=1, row=2, pady=10, padx=5)

button_converter.grid(column=2, row=2, pady=10, padx=5)

label_gte.grid(column=4, row=2, padx=5, sticky=E)

top_frame2.grid_columnconfigure(2, weight=4)

button_exit.grid(column=3, row=7, padx=5)
low_frame.grid_columnconfigure(4, weight=4)
label_auth.grid(column=5, row=7, padx=5, sticky=E)

# Let the window wait for any events
window.mainloop()
