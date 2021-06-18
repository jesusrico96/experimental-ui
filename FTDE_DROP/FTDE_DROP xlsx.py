from tkinter import *
import os
import os.path
import sys
import matplotlib.pyplot as plt
import statistics
import seaborn as sns
import pandas as pd
import openpyxl as opx
from PIL import *
from tkinter import filedialog

global M
global SD
global median
global MAD
M = 0
SD = 0
median = 0
MAD = 0


def browseFiles():
    global filename
    filename = filedialog.askopenfilename(initialdir=sys.path[0],
                                          title="Escolle un arquivo",
                                          filetypes=(("Ficheiros de Excel",
                                                      "*.xlsx"),
                                                     ("Todos os formatos",
                                                      "*.*")))

    # Etiqueta do ficheiro utilizado
    label_file_explorer.configure(text=filename)

    my_file = opx.load_workbook(filename=filename)
    my_sheet = my_file.worksheets[0]
    global wb
    global datos
    wb = opx.Workbook()
    datos = wb.active
    mr = my_sheet.max_row
    mc = my_sheet.max_column
    for i in range(1, mr + 1):
        for j in range(1, mc + 1):
            c = my_sheet.cell(row=i, column=j)
            datos.cell(row=i, column=j).value = c.value


def converter():
    # Na carpeta do script, lista os arquivos que contén:
    cwd = os.path.abspath('')
    files = os.listdir(cwd)

    for file in files:
        if file.endswith('.lvm'):
            my_file = open(file)
            string_list = my_file.readlines()
            string_list[0:21] = ""  # Crea 0 liñas baleiras no espazo ocupado antes por 21 liñas.
            my_file.close()

            my_file = open(file, "w")
            new_file_contents = "".join(string_list)
            my_file.write(new_file_contents)
            my_file.close()

            my_file = open(file)
            df = pd.read_csv(my_file, sep="\t", decimal=',', header=0)
            file2 = file.replace('.lvm', '')
            df.to_excel(file2+".xlsx")
            my_file.close()

            my_file = opx.load_workbook(filename=file2+".xlsx")
            my_sheet = my_file.worksheets[0]
            my_sheet.delete_cols(28, 8)
            my_sheet.delete_cols(20)
            my_sheet.delete_cols(12, 3)
            my_sheet.delete_cols(1, 2)

            # Contar filas e columnas do arquivo
            mr = my_sheet.max_row

            # Corrección de gases
            my_sheet.cell(1, 23).value = "NO (ppm) corr"
            for i in range(2, mr + 1):
                ttl = my_sheet.cell(i, 10).value * ((21-10)/(21-my_sheet.cell(i, 12).value))
                my_sheet.cell(i, 23).value = ttl
                # my_sheet.move_range("W1:W"+str(mr), cols=-13)
            print("NO corrected!")

            my_sheet.cell(1, 24).value = "NOx (ppm) corr"
            for i in range(2, mr + 1):
                ttl = my_sheet.cell(i, 11).value * ((21 - 10) / (21 - my_sheet.cell(i, 12).value))
                my_sheet.cell(i, 24).value = ttl
                # my_sheet.move_range("X:X", cols=-13)
            print("NOx corrected!")

            my_sheet.cell(1, 25).value = "CO (ppm) corr"
            for i in range(2, mr + 1):
                ttl = my_sheet.cell(i, 13).value * ((21 - 10) / (21 - my_sheet.cell(i, 12).value))*10000
                my_sheet.cell(i, 25).value = ttl
                # my_sheet.move_range("Y:Y", cols=-12)
            print("CO corrected!")

            my_sheet.cell(1, 26).value = "CO2 (%) corr"
            for i in range(2, mr + 1):
                ttl = my_sheet.cell(i, 14).value * ((21 - 10) / (21 - my_sheet.cell(i, 12).value))
                my_sheet.cell(i, 26).value = ttl
                # my_sheet.move_range("Z:Z", cols=-12)
            print("CO2 corrected!")

            my_sheet.delete_cols(13, 2)
            my_sheet.delete_cols(10, 2)

            my_file.save(filename=file2+".xlsx")
    print("Files converted!")


# Fixación tempos
def time():
    global Ti
    global Tf
    try:
        Ti = float(Tempiniw.get())
        Tf = float(Tempfinw.get())
        print("Time set to interval between " + str(Ti) + " and " + str(Tf) + " s.")

    except:
        Ti = "tiempo1"
        Tf = "tiempo2"
        print("Time restrictions removed.")


# Media e desv. típica no rango de tempos
def MDT():
    try:
        # Tempo:
        tempo = list(datos.columns)[1]  # Esto é unha tupla que contén cellObj
        # Lista que contén todos os tempos en orde:
        tempo2 = []
        # Lista que contén todos os tempos solicitados:
        tempo3 = []
        # Lista que contén todos os índices dos tempos solicitados na lista global,
        # de xeito que se podan relacionar con outros parámetros:
        tindex = []
        for cellObj in tempo:
            x = cellObj.value
            tempo2.append(x)
        tempo2.pop(0)
        for x in tempo2:
            if Tf > x > Ti:
                tempo3.append(x)
                tindex.append(tempo2.index(x))
            elif x > Tf:
                break

        # Variable a analizar
        analito = list(datos.columns)[opcionso1.index(variableo1.get())]
        analito2 = []
        analito3 = []
        for cellObj in analito:
            x = cellObj.value
            analito2.append(x)
        analito2.pop(0)

        for x in analito2:
            for n in tindex:
                if analito2.index(x) == n and len(analito3) < len(tindex):
                    analito3.append(x)
                elif len(analito3) == len(tindex):
                    break
                elif x is None:
                    break

        M = round(statistics.mean(analito3), 2)
        median = round(statistics.median(analito3), 2)
        SD = round(statistics.stdev(analito3), 2)
        series = pd.Series(analito3)
        MAD = round(series.mad())

    except:
        # Tempo:
        tempo = list(datos.columns)[1]  # Esto é unha tupla que contén cellObj
        tempo2 = []  # Esta lista conterá todos os tempos en orde
        for cellObj in tempo:
            x = cellObj.value
            tempo2.append(x)
        tempo2.pop(0)

        # Variable a analizar
        analito = list(datos.columns)[opcionso1.index(variableo1.get())]
        analito2 = []
        for cellObj in analito:
            x = cellObj.value
            analito2.append(x)
        analito2.pop(0)

        M = round(statistics.mean(analito2), 2)
        median = round(statistics.median(analito2), 2)
        SD = round(statistics.stdev(analito2), 2)
        series = pd.Series(analito2)
        MAD = round(series.mad())

    label_Mean.configure(text="Mean" + " " + str(M))
    label_Median.configure(text="Median:" + " " + str(median))
    label_SD.configure(text="Desv. Est." + " " + str(SD))
    label_MAD.configure(text="Mean Abs. Dev.:" + " " + str(MAD))


# Gráficos
def grafico():
    # Se a checkbox está desmarcada, fai gráficos só cunha variable en Y. Dentro de cada tipo de gráfico
    # hai un bucle para comprobar se se pide o gráfico nun tempo limitado ou para todo o set de datos.
    if bo2.get() == 0:
        try:
            # Tempo:
            # Esto é unha tupla que contén cellObj:
            abscisa = list(datos.columns)[opcionsa1.index(variablea1.get())]
            # Esta lista conterá todos os tempos en orde:
            abscisa2 = []
            # Esta lista conterá só os tempos solicitados:
            abscisa3 = []
            # Lista que contén todos os índices dos tempos solicitados na lista global,
            # de xeito que se podan relacionar con outros parámetros:
            tindex = []
            for cellObj in abscisa:
                x = cellObj.value
                abscisa2.append(x)
            abscisa2.pop(0)
            for x in abscisa2:
                if Tf > x > Ti:
                    abscisa3.append(x)
                    tindex.append(abscisa2.index(x))
                elif x > Tf:
                    break

            # Variable a analizar
            analito = list(datos.columns)[opcionso1.index(variableo1.get())]
            analito2 = []
            analito3 = []
            for cellObj in analito:
                x = cellObj.value
                analito2.append(x)
            analito2.pop(0)

            for x in analito2:
                for n in tindex:
                    if analito2.index(x) == n and len(analito3) < len(tindex):
                        analito3.append(x)
                    elif len(analito3) == len(tindex):
                        break

            # plot
            plt.plot(abscisa3, analito3, label=variableo1.get(), linewidth=0,
                     marker='.', markerfacecolor='blue', markersize=3)

            # naming the x axis
            plt.xlabel(variablea1.get())
            # naming the y axis
            plt.ylabel(variableo1.get())
            # Legend
            plt.legend()

            # giving a title to my graph
            plt.title(str(variablea1.get()) + " vs. " + str(variableo1.get()))

        except:
            # Tempo:
            abscisa = list(datos.columns)[opcionsa1.index(variablea1.get())]  # Esto é unha tupla que contén cellObj
            abscisa2 = []  # Esta lista conterá todos os tempos en orde
            for cellObj in abscisa:
                x = cellObj.value
                abscisa2.append(x)
            abscisa2.pop(0)

            # Variable a analizar
            analito = list(datos.columns)[opcionso1.index(variableo1.get())]
            analito2 = []
            for cellObj in analito:
                x = cellObj.value
                analito2.append(x)
            analito2.pop(0)

            # plot
            plt.plot(abscisa2, analito2, label=variableo1.get(), linewidth=0,
                     marker='.', markerfacecolor='blue', markersize=3)

            plt.xlabel(variablea1.get())
            plt.ylabel(variableo1.get())
            plt.legend()

            plt.title(str(variablea1.get()) + " vs. " + str(variableo1.get()))
    else:
        try:
            # Tempo:
            abscisa = list(datos.columns)[opcionsa1.index(variablea1.get())]
            abscisa2 = []
            abscisa3 = []
            tindex = []
            for cellObj in abscisa:
                x = cellObj.value
                abscisa2.append(x)
            abscisa2.pop(0)
            for x in abscisa2:
                if Tf > x > Ti:
                    abscisa3.append(x)
                    tindex.append(abscisa2.index(x))
                elif x > Tf:
                    break

            # Variables a analizar
            analitoa = list(datos.columns)[opcionso1.index(variableo1.get())]
            analitoa2 = []
            analitoa3 = []
            for cellObj in analitoa:
                x = cellObj.value
                analitoa2.append(x)
            analitoa2.pop(0)

            for x in analitoa2:
                for n in tindex:
                    if analitoa2.index(x) == n and len(analitoa3) < len(tindex):
                        analitoa3.append(x)
                    elif len(analitoa3) == len(tindex):
                        break

            analitob = list(datos.columns)[opcionso2.index(variableo2.get())]
            analitob2 = []
            analitob3 = []
            for cellObj in analitob:
                x = cellObj.value
                analitob2.append(x)
            analitob2.pop(0)

            for x in analitob2:
                for n in tindex:
                    if analitob2.index(x) == n and len(analitob3) < len(tindex):
                        analitob3.append(x)
                    elif len(analitob3) == len(tindex):
                        break

            # plot
            plt.plot(abscisa3, analitoa3, label=variableo1.get(), linewidth=0,
                     marker='.', markerfacecolor='blue', markersize=3)

            plt.plot(abscisa3, analitob3, label=variableo2.get(), linewidth=0,
                     marker='.', markerfacecolor='r', markersize=3)

            plt.xlabel(variablea1.get())
            plt.ylabel(variableo1.get() + " and " + variableo2.get())
            plt.legend()

            plt.title(str(variablea1.get()) + " vs. " + str(variableo1.get()) + " and " + variableo2.get())

        except:
            # Tempo:
            abscisa = list(datos.columns)[opcionsa1.index(variablea1.get())]  # Esto é unha tupla que contén cellObj
            abscisa2 = []  # Esta lista conterá todos os tempos en orde
            for cellObj in abscisa:
                x = cellObj.value
                abscisa2.append(x)
            abscisa2.pop(0)

            # Variables a analizar
            analitoa = list(datos.columns)[opcionso1.index(variableo1.get())]
            analitoa2 = []
            for cellObj in analitoa:
                x = cellObj.value
                analitoa2.append(x)
            analitoa2.pop(0)

            analitob = list(datos.columns)[opcionso2.index(variableo2.get())]
            analitob2 = []
            for cellObj in analitob:
                x = cellObj.value
                analitob2.append(x)
            analitob2.pop(0)

            # plot
            plt.plot(abscisa2, analitoa2, label=variableo1.get(), linewidth=0,
                     marker='.', markerfacecolor='blue', markersize=3)

            plt.plot(abscisa2, analitob2, label=variableo2.get(), linewidth=0,
                     marker='.', markerfacecolor='r', markersize=3)

            plt.xlabel(variablea1.get())
            plt.ylabel(variableo1.get() + " and " + variableo2.get())
            plt.legend()

            plt.title(str(variablea1.get()) + " vs. " + str(variableo1.get()) + " and " + variableo2.get())

    # function to show the plot
    plt.show()


# Boxplots
def boxplot():
    if bv3.get() == 0:
        print("1 independent variable detected, drawing 1 Boxplot.")
        try:
            # Tempo:
            abscisa = list(datos.columns)[opcionsv1.index(variable1.get())]
            abscisa2 = []
            abscisa3 = []
            tindex = []
            for cellObj in abscisa:
                x = cellObj.value
                abscisa2.append(x)
            abscisa2.pop(0)

            print("Applying temporal restraint to X Axis")
            for x in abscisa2:
                if Tf > x > Ti:
                    abscisa3.append(x)
                    tindex.append(abscisa2.index(x))
                elif x > Tf:
                    break
            print("Temporal limitation to X Axis successful")

            # Variable a analizar
            analito = list(datos.columns)[opcionsv2.index(variable2.get())]
            analito2 = []
            analito3 = []
            for cellObj in analito:
                x = cellObj.value
                analito2.append(x)
            analito2.pop(0)

            print("Applying temporal restraint to Y Axis")
            for x in analito2:
                for n in tindex:
                    if analito2.index(x) == n and len(analito3) < len(tindex):
                        analito3.append(x)
                    elif len(analito3) == len(tindex):
                        break
            print("Temporal limitation to Y Axis successful")

            sns.boxplot(data=analito3)

            # naming the axis
            plt.xlabel(variable2.get())

        except:
            print("No temporal restriction detected.")

            # Variable a analizar
            analito = list(datos.columns)[opcionsv2.index(variable2.get())]
            analito2 = []
            for cellObj in analito:
                x = cellObj.value
                analito2.append(x)
            analito2.pop(0)
            print("Variable 2 data extracted.")

            sns.boxplot(data=analito2)

            # naming the axes

            plt.xlabel(variable2.get())
    else:
        print("2 independent variables detected, drawing 2 Boxplots.")
        try:
            # Tempo:
            abscisa = list(datos.columns)[opcionsv1.index(variable1.get())]
            abscisa2 = []
            abscisa3 = []
            tindex = []
            for cellObj in abscisa:
                x = cellObj.value
                abscisa2.append(x)
            abscisa2.pop(0)
            print("Applying temporal restraint to X Axis")
            for x in abscisa2:
                if Tf > x > Ti:
                    abscisa3.append(x)
                    tindex.append(abscisa2.index(x))
                elif x > Tf:
                    break
            print("Temporal limitation to X Axis successful")

            # Variables a analizar
            analitoa = list(datos.columns)[opcionsv2.index(variable2.get())]
            analitoa2 = []
            analitoa3 = []
            for cellObj in analitoa:
                x = cellObj.value
                analitoa2.append(x)
            analitoa2.pop(0)

            print("Applying temporal limitation to Y1 Axis.")
            for x in analitoa2:
                for n in tindex:
                    if analitoa2.index(x) == n and len(analitoa3) < len(tindex):
                        analitoa3.append(x)
                    elif len(analitoa3) == len(tindex):
                        break
            print("Temporal limitation to Y1 Axis successful")

            analitob = list(datos.columns)[opcionsv3.index(variable3.get())]
            analitob2 = []
            analitob3 = []
            for cellObj in analitob:
                x = cellObj.value
                analitob2.append(x)
            analitob2.pop(0)

            print("Applying temporal limitation to Y2 Axis.")
            for x in analitob2:
                for n in tindex:
                    if analitob2.index(x) == n and len(analitob3) < len(tindex):
                        analitob3.append(x)
                    elif len(analitob3) == len(tindex):
                        break
            print("Temporal limitation to Y2 Axis successful")

            df = pd.DataFrame(list(zip(analitoa3, analitob3)))
            df.columns = [str(variable2.get()), str(variable3.get())]

            sns.boxplot(data=df)

            # naming the y axis
            plt.ylabel(variable2.get() + " and " + variable3.get())

        except:
            print("No temporal restriction detected.")

            # Variables a analizar
            analitoa = list(datos.columns)[opcionsv2.index(variable2.get())]
            analitoa2 = []
            for cellObj in analitoa:
                x = cellObj.value
                analitoa2.append(x)
            analitoa2.pop(0)
            print("Variable 2 data acquired.")

            analitob = list(datos.columns)[opcionsv3.index(variable3.get())]
            analitob2 = []
            for cellObj in analitob:
                x = cellObj.value
                analitob2.append(x)
            analitob2.pop(0)
            print("Variable 3 data obtained.")

            df = pd.DataFrame(list(zip(analitoa2, analitob2)))
            df.columns = [str(variable2.get()), str(variable3.get())]

            sns.boxplot(data=df)

            # naming the y axis
            plt.ylabel(variable2.get() + " and " + variable3.get())
    # function to show the plot
    plt.show()


# Ventana principal
window = Tk()

# Titulo da ventá principal
window.title('FTDE: DROP')

# Tamaño da ventá principal
window.geometry("725x450")

# Cor de fondo da venta principal
window.config(background="white")

# Imos crear algúns frames de fondo
top_frame = Frame(window, bg="black", width=450, height=50, pady=3)
top_frame2 = Frame(window, bg="alice blue", width=450, height=50, pady=6)
middle_frame = Frame(window, bg="black", width=450, height=400, pady=3)
low_frame = Frame(window, bg="gray84", width=450, height=50, pady=3)

# Disposición dos frames de fondo
window.grid_rowconfigure(1, weight=1)
window.grid_columnconfigure(0, weight=1)

top_frame.grid(row=0, sticky="ew")
top_frame2.grid(row=1, sticky="nsew")
middle_frame.grid(row=2, sticky="ew")
low_frame.grid(row=3, sticky="ew")

# Creación e disposición dos frames de fronte (centro)
middle_frame.grid_rowconfigure(0, weight=1)
middle_frame.grid_columnconfigure(1, weight=1)

framea = Frame(middle_frame, bg='alice blue', width=250, height=400, padx=3, pady=3)
frameb = Frame(middle_frame, bg='alice blue', width=250, height=400, padx=3, pady=3)

framea.grid(row=0, column=0, columnspan=2, sticky="ns", padx=3)
frameb.grid(row=0, column=2, sticky="ns", padx=3)

# Etiquetas do Explorador de Arquivos
label_file_explorer = Label(top_frame,
                            text="Ferramenta de tratamento de datos experimentais do GTE: versión DROP",
                            width=105, height=4,
                            fg="blue")

# Etiqueta co logo do GTE
dirname = os.path.dirname(__file__)
photogte = PhotoImage(file=os.path.join(dirname, "imaxes", "gte2.png"))

photoimagegte = photogte.subsample(3, 3)

label_gte = Label(top_frame2,
                  image=photoimagegte)

# Botóns do Explorador de Arquivos
button_explore = Button(top_frame2,
                        text="Navegar arquivos",
                        command=browseFiles)

button_exit = Button(low_frame,
                     text="Saír",
                     fg="red",
                     command=exit)

# Etiqueta autoría
label_auth = Label(low_frame,
                   text="GNU-GPL   2021   Juan Jesús Rico Fuentes",
                   fg="black")

# Botón de guillotina
photo = PhotoImage(
    file=os.path.join(dirname, "imaxes", "conv.png"))

photoimage = photo.subsample(25, 25)

button_converter = Button(top_frame2,
                           text="Conversor",
                           image=photoimage,
                           command=converter)

# Widget entrada de tempos:

# Etiqueta para tempo inicial
Tempini_label = Label(top_frame2,
                      text='Tempo inicial')

# Entrada do tempo incial
Tempiniw = Entry(top_frame2,
                 width=10)

# Etiqueta para tempo final
Tempfin_label = Label(top_frame2,
                      text='Tempo final')

# Entrada do tempo final
Tempfinw = Entry(top_frame2,
                 width=10)

# Fixar tempos
button_tempos = Button(top_frame2,
                       text="Fixar tempos",
                       command=time)

# Etiqueta gráficos
label_partial_graph = Label(framea,
                            width=70, height=1,
                            text="Gráficos")

# Media e desv. típ.
button_MDT = Button(framea,
                    text="MDT (y)",
                    command=MDT)

# Etiqueta coa media
label_Mean = Label(framea,
                   text="Media:" + " " + str(M),
                   fg="blue")

# Etiqueta coa mediana
label_Median = Label(framea,
                     text="Median:" + " " + str(median),
                     fg="blue")

# Etiqueta coa standard dev.
label_SD = Label(framea,
                 text="Std. Dev.:" + " " + str(SD),
                 fg="blue")

# Etiqueta coa MAD
label_MAD = Label(framea,
                  text="Mean Abs. Dev.:" + " " + str(MAD),
                  fg="blue")

# Botón de Gráficos
button_graph = Button(framea,
                      text="Gráfico",
                      width=10, height=2,
                      command=grafico)

# Botón de Boxplots
button_boxp1 = Button(frameb,
                      text="Box Plot 1",
                      width=10, height=2,
                      command=boxplot)

# Etiqueta de boxplots
label_global_graph = Label(frameb,
                           width=20, height=1,
                           text="Boxplots")

# Etiqueta "Eixo X" grafico
label_a1 = Label(framea,
                 text="Eixo X:")

# Etiqueta "Eixo Y" grafico
label_o1 = Label(framea,
                 text="Eixo Y:")

# Etiqueta "Eixo Y 2" grafico e checkbutton para activala
label_o2 = Label(framea,
                 text="Eixo Y 2:")

bo2 = IntVar()
button_o2 = Checkbutton(framea, text='', variable=bo2, onvalue=1, offvalue=0)

# Etiqueta "Var 1" boxplot
label_v1 = Label(frameb,
                 text="Variable 1 (t):")

# Etiqueta "Var 2" boxplot
label_v2 = Label(frameb,
                 text="Variable 2:")

# Etiqueta "Var 3" boxplot con botón para activar
label_v3 = Label(frameb,
                 text="Variable 3:")
bv3 = IntVar()
button_v3 = Checkbutton(frameb, text='', variable=bv3, onvalue=1, offvalue=0)

# Widget para selección de variable a plottear. A orde das variables é a do LabView do Drop, así que haberá que
# axustala se se cambia de instalación

# Lista e widget de selección para abscisas do gráfico
opcionsa1 = ["Segundos", "Tª Agua Entrada", "Tª Agua Salida", "Caudal Aire Prim", "Cauda Aire Sec",
             "Tª Humos", "Tª Chimenea", "Caudal Agua", "Peso Pellet (kg)", "O2 (%)", "TC1", "TC2",
             "TC3", "TC4", "TC5", "TC6", "TC7", "Comment", "NO (ppm) corr", "NOx (ppm) corr",
             "CO (ppm) corr", "CO2 (%) corr"]

variablea1 = StringVar(window)
variablea1.set(opcionsa1[0])

menua1 = OptionMenu(framea, variablea1, *opcionsa1)

# Lista e widget de selección para ordenadas 1 do gráfico
opcionso1 = ["Segundos", "Tª Agua Entrada", "Tª Agua Salida", "Caudal Aire Prim", "Cauda Aire Sec",
             "Tª Humos", "Tª Chimenea", "Caudal Agua", "Peso Pellet (kg)", "O2 (%)", "TC1", "TC2",
             "TC3", "TC4", "TC5", "TC6", "TC7", "Comment", "NO (ppm) corr", "NOx (ppm) corr",
             "CO (ppm) corr", "CO2 (%) corr"]

variableo1 = StringVar(window)
variableo1.set(opcionso1[1])

menuo1 = OptionMenu(framea, variableo1, *opcionso1)

# Lista e widget de selección para ordenadas 2 do gráfico
opcionso2 = ["Segundos", "Tª Agua Entrada", "Tª Agua Salida", "Caudal Aire Prim", "Cauda Aire Sec",
             "Tª Humos", "Tª Chimenea", "Caudal Agua", "Peso Pellet (kg)", "O2 (%)", "TC1", "TC2",
             "TC3", "TC4", "TC5", "TC6", "TC7", "Comment", "NO (ppm) corr", "NOx (ppm) corr",
             "CO (ppm) corr", "CO2 (%) corr"]

variableo2 = StringVar(window)
variableo2.set(opcionso2[2])

menuo2 = OptionMenu(framea, variableo2, *opcionso2)

# Lista e widget de selección para Variable 1 Boxplot
opcionsv1 = ["Segundos", "Tª Agua Entrada", "Tª Agua Salida", "Caudal Aire Prim", "Cauda Aire Sec",
             "Tª Humos", "Tª Chimenea", "Caudal Agua", "Peso Pellet (kg)", "O2 (%)", "TC1", "TC2",
             "TC3", "TC4", "TC5", "TC6", "TC7", "Comment", "NO (ppm) corr", "NOx (ppm) corr",
             "CO (ppm) corr", "CO2 (%) corr"]

variable1 = StringVar(window)
variable1.set(opcionsv1[0])

menuv1 = OptionMenu(frameb, variable1, *opcionsv1)

# Lista e widget de selección para Variable 2 Boxplot
opcionsv2 = ["Segundos", "Tª Agua Entrada", "Tª Agua Salida", "Caudal Aire Prim", "Cauda Aire Sec",
             "Tª Humos", "Tª Chimenea", "Caudal Agua", "Peso Pellet (kg)", "O2 (%)", "TC1", "TC2",
             "TC3", "TC4", "TC5", "TC6", "TC7", "Comment", "NO (ppm) corr", "NOx (ppm) corr",
             "CO (ppm) corr", "CO2 (%) corr"]

variable2 = StringVar(window)
variable2.set(opcionsv2[1])

menuv2 = OptionMenu(frameb, variable2, *opcionsv2)

# Lista e widget de selección para Variable 3 Boxplot
opcionsv3 = ["Segundos", "Tª Agua Entrada", "Tª Agua Salida", "Caudal Aire Prim", "Cauda Aire Sec",
             "Tª Humos", "Tª Chimenea", "Caudal Agua", "Peso Pellet (kg)", "O2 (%)", "TC1", "TC2",
             "TC3", "TC4", "TC5", "TC6", "TC7", "Comment", "NO (ppm) corr", "NOx (ppm) corr",
             "CO (ppm) corr", "CO2 (%) corr"]

variable3 = StringVar(window)
variable3.set(opcionsv3[2])

menuv3 = OptionMenu(frameb, variable3, *opcionsv3)

# Grid method is chosen for placing
# the widgets at respective positions
# in a table like structure by
# specifying rows and columns

label_file_explorer.grid(column=1, row=1)

button_converter.grid(column=1, row=1, rowspan=2, pady=10, padx=5)
button_explore.grid(column=2, row=1, rowspan=2, pady=10, padx=5)

label_gte.grid(column=8, row=1, rowspan=2, padx=5)

top_frame2.grid_columnconfigure(7, weight=4)

button_exit.grid(column=3, row=7, padx=5)
low_frame.grid_columnconfigure(4, weight=4)
label_auth.grid(column=5, row=7, padx=5, sticky=E)

# Widget entrada de tempos


Tempini_label.grid(column=3, row=1, pady=5, padx=5)
Tempiniw.grid(column=4, row=1, pady=5, padx=5)
Tempfin_label.grid(column=3, row=2, pady=5, padx=5)
Tempfinw.grid(column=4, row=2, pady=5, padx=5)
button_tempos.grid(column=5, row=1, rowspan=2, pady=5, padx=5)

# Gráficos
label_partial_graph.grid(column=1, row=1, columnspan=6, pady=5, padx=5)

button_MDT.grid(column=4, row=2, rowspan=4, pady=5, padx=5)
label_Mean.grid(column=5, row=2, pady=5, padx=5)
label_Median.grid(column=5, row=3, pady=5, padx=5)
label_SD.grid(column=5, row=4, pady=5, padx=5)
label_MAD.grid(column=5, row=5, pady=5, padx=5)

# Widget de lista desplegable abscisas

label_a1.grid(column=2, row=2, pady=5, padx=5, sticky=E)
menua1.grid(column=3, row=2, pady=5, padx=5)

# Widget de lista desplegable ordenadas 1

label_o1.grid(column=2, row=3, pady=5, padx=5, sticky=E)
menuo1.grid(column=3, row=3, pady=5, padx=5)

# Widget de lista desplegable ordenadas 2

label_o2.grid(column=2, row=4, pady=5, padx=5, sticky=E)
menuo2.grid(column=3, row=4, pady=5, padx=5)
button_o2.grid(column=2, row=4, pady=5, sticky=W)

button_graph.grid(column=1, row=5, columnspan=3, pady=5)

# Widget de Variable temporal Boxplot

label_v1.grid(column=2, row=1, pady=5, padx=5, sticky=E)
label_global_graph.grid(column=2, row=0, columnspan=2, pady=5, padx=5, sticky="N")
menuv1.grid(column=3, row=1, pady=8, padx=5)
frameb.rowconfigure((1, 5), weight=1)

# Widget de Variable 2 Boxplot

label_v2.grid(column=2, row=3, pady=5, padx=5, sticky=E)
menuv2.grid(column=3, row=3, pady=8, padx=5)

# Widget de Variable 3 Boxplot + Checkbutton

label_v3.grid(column=2, row=4, pady=5, padx=5, sticky=E)
button_v3.grid(column=2, row=4, pady=5, sticky=W)
menuv3.grid(column=3, row=4, pady=5, padx=5)

button_boxp1.grid(column=1, row=5, columnspan=3, pady=5, padx=5)

# Let the window wait for any events
window.mainloop()
