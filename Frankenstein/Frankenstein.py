import datetime
from tkinter import *
from openpyxl import *
from PIL import *
import os
import pandas as pd
from datetime import *
# Explorador de arquivos
from tkinter import filedialog

filename1 = 0
filename2 = 0
filename3 = 0
filename4 = 0
filename5 = 0


def browseFiles1():
    global filename1
    filename1 = filedialog.askopenfilename(initialdir=sys.path[0],
                                           title="Escolle un arquivo inicial",
                                           filetypes=(("Ficheiro de Excel",
                                                       "*.xlsx"),
                                                      ("Datos de LabView",
                                                       "*.lvm*"),
                                                      ("Ficheiros de texto",
                                                       "*.txt*"),
                                                      ("Todos os formatos",
                                                       "*.*")))

    # Etiqueta do ficheiro utilizado
    label_file_explorer.configure(text=filename1)
    print(filename1)


def browseFiles2():
    global filename2
    filename2 = filedialog.askopenfilename(initialdir=sys.path[0],
                                           title="Escolle un arquivo 2",
                                           filetypes=(("Ficheiro de Excel",
                                                       "*.xlsx"),
                                                      ("Datos de LabView",
                                                       "*.lvm*"),
                                                      ("Ficheiros de texto",
                                                       "*.txt*"),
                                                      ("Todos os formatos",
                                                       "*.*")))


def browseFiles3():
    global filename3
    filename3 = filedialog.askopenfilename(initialdir=sys.path[0],
                                           title="Escolle un arquivo 3",
                                           filetypes=(("Ficheiro de Excel",
                                                       "*.xlsx"),
                                                      ("Datos de LabView",
                                                       "*.lvm*"),
                                                      ("Ficheiros de texto",
                                                       "*.txt*"),
                                                      ("Todos os formatos",
                                                       "*.*")))


def browseFiles4():
    global filename4
    filename4 = filedialog.askopenfilename(initialdir=sys.path[0],
                                           title="Escolle un arquivo 4",
                                           filetypes=(("Ficheiro de Excel",
                                                       "*.xlsx"),
                                                      ("Datos de LabView",
                                                       "*.lvm*"),
                                                      ("Ficheiros de texto",
                                                       "*.txt*"),
                                                      ("Todos os formatos",
                                                       "*.*")))


def browseFiles5():
    global filename5
    filename5 = filedialog.askopenfilename(initialdir=sys.path[0],
                                           title="Escolle un arquivo 5",
                                           filetypes=(("Ficheiro de Excel",
                                                       "*.xlsx"),
                                                      ("Datos de LabView",
                                                       "*.lvm*"),
                                                      ("Ficheiros de texto",
                                                       "*.txt*"),
                                                      ("Todos os formatos",
                                                       "*.*")))


def frankenstein():
    # Abrir o arquivo 1 cos datos e un workbook paralelo de destino dos mesmos
    my_file = load_workbook(filename=filename1)
    my_sheet = my_file.worksheets[1]
    global wb
    global datos
    wb = Workbook()
    datos = wb.active

    # Contar filas e columnas do arquivo orixinal e pegalas no arquivo de destino
    mr = my_sheet.max_row
    mc = my_sheet.max_column

    for i in range(1, mr + 1):
        for j in range(1, mc + 1):
            c = my_sheet.cell(row=i, column=j)
            datos.cell(row=i, column=j).value = c.value

    wb.save("File1.xls")

    # Abrir o arquivo 2 cos datos e un workbook paralelo de destino dos mesmos
    my_file2 = load_workbook(filename=filename2)
    my_sheet2 = my_file2.worksheets[1]
    global wb2
    global datos2
    wb2 = Workbook()
    datos2 = wb2.active

    # Contar filas e columnas do arquivo orixinal e pegalas no arquivo de destino
    mr2 = my_sheet2.max_row
    mc2 = my_sheet2.max_column

    for i in range(1, mr2 + 1):
        for j in range(1, mc2 + 1):
            c2 = my_sheet2.cell(row=i, column=j)
            datos2.cell(row=i, column=j).value = c2.value

    wb2.save("File2.xls")

    # Abrir o arquivo 3 cos datos e un workbook paralelo de destino dos mesmos
    if filename3 != 0:
        my_file3 = load_workbook(filename=filename3)
        my_sheet3 = my_file3.worksheets[1]
        global wb3
        global datos3
        wb3 = Workbook()
        datos3 = wb3.active

        # Contar filas e columnas do arquivo orixinal e pegalas no arquivo de destino
        mr3 = my_sheet3.max_row
        mc3 = my_sheet3.max_column

        for i in range(1, mr3 + 1):
            for j in range(1, mc3 + 1):
                c3 = my_sheet3.cell(row=i, column=j)
                datos3.cell(row=i, column=j).value = c3.value
        wb3.save("File3.xls")
    else:
        wb3 = Workbook()
        wb3.save("File3.xls")

    # Abrir o arquivo 4 cos datos e un workbook paralelo de destino dos mesmos
    if filename4 != 0:
        my_file4 = load_workbook(filename=filename4)
        my_sheet4 = my_file4.worksheets[1]
        global wb4
        global datos4
        wb4 = Workbook()
        datos4 = wb3.active

        # Contar filas e columnas do arquivo orixinal e pegalas no arquivo de destino
        mr4 = my_sheet4.max_row
        mc4 = my_sheet4.max_column

        for i in range(1, mr4 + 1):
            for j in range(1, mc4 + 1):
                c4 = my_sheet4.cell(row=i, column=j)
                datos4.cell(row=i, column=j).value = c4.value
        wb4.save("File4.xls")
    else:
        wb4 = Workbook()
        wb4.save("File4.xls")

    # Abrir o arquivo 5 cos datos e un workbook paralelo de destino dos mesmos
    if filename5 != 0:
        my_file5 = load_workbook(filename=filename5)
        my_sheet5 = my_file5.worksheets[1]
        global wb5
        global datos5
        wb5 = Workbook()
        datos5 = wb5.active

        # Contar filas e columnas do arquivo orixinal e pegalas no arquivo de destino
        mr5 = my_sheet5.max_row
        mc5 = my_sheet5.max_column

        for i in range(1, mr5 + 1):
            for j in range(1, mc5 + 1):
                c5 = my_sheet5.cell(row=i, column=j)
                datos5.cell(row=i, column=j).value = c5.value
        wb5.save("File5.xls")
    else:
        wb5 = Workbook()
        wb5.save("File5.xls")

    # Na carpeta do script, lista os arquivos que contén:
    cwd = os.path.abspath('')
    files = os.listdir(cwd)

    # Dos arquivos listados, une os .xls que se crearon antes entre si:
    df = pd.DataFrame()
    for file in files:
        if file.endswith('.xls'):
            df = df.append(pd.read_excel(file), ignore_index=True)

    # Finalmente, escribe un .xlsx con todos os datos:
    df.to_excel("Frankenstein.xlsx", index=False)

    # Conversión dos tempos a un marco de referencia global
    # Primeiro, cárgase o arquivo xerado
    frk = load_workbook(filename="Frankenstein.xlsx")
    frksh = frk.worksheets[0]
    # print(frksh.cell(2,1).value)

    mrfrk = frksh.max_row

    # Prepárase un bucle que reste dous valores de tempo e lles sume o valor anterior a eles
    # Deste xeito úsase a timestamp de cada dato para construir o marco temporal global
    frksh.cell(1, 69).value = "Tiempo corr."
    frksh.cell(2, 69).value = 1.0
    frksh.cell(3, 69).value = 2.0
    for i in range(4, mrfrk + 1):
        t1 = datetime.strptime(frksh.cell(i, 1).value, "%H:%M:%S")
        t2 = datetime.strptime(frksh.cell(i - 1, 1).value, "%H:%M:%S")
        t3 = frksh.cell(i - 1, 69).value
        seg = t1 - t2
        seg.seconds
        seg2 = t3 + (seg / pd.to_timedelta(1, unit="S"))
        frksh.cell(i, 69).value = seg2

    # Móvese o marco global á columna axeitada
    frksh.move_range("BQ2:BQ7500", cols=-67)

    # Pechar os arquivos orixinais e gardar un Excel cos datos de traballo
    frk.save(filename="Frankenstein.xlsx")
    frk.close()
    my_file.close()
    my_file2.close()
    if filename3 != 0:
        my_file3.close()
    if filename4 != 0:
        my_file4.close()
    if filename5 != 0:
        my_file5.close()

    print("Patchwork done!")  # Amosa unha confirmación da execución en consola.


# Ventana principal
window = Tk()

# Titulo da ventá principal
window.title('Frankenstein: BICO')

# Tamaño da ventá principal
window.geometry("900x240")

# Cor de fondo da venta principal
window.config(background="white")

# Imos crear algúns frames de fondo
top_frame = Frame(window, bg="black", width=450, height=50, pady=3)
top_frame2 = Frame(window, bg="alice blue", width=450, height=50, pady=6)
low_frame = Frame(window, bg="gray84", width=450, height=50, pady=3)

# Disposición dos frames de fondo
window.grid_rowconfigure(1, weight=1)
window.grid_columnconfigure(0, weight=1)

top_frame.grid(row=0, sticky="ew")
top_frame2.grid(row=1, sticky="ew")
low_frame.grid(row=3, sticky="ew")

# Etiquetas do Explorador de Arquivos
label_file_explorer = Label(top_frame,
                            text="Ferramenta de unión de datos fragmentados do GTE: versión BICO",
                            width=130, height=4,
                            fg="blue")

label_oblig = Label(top_frame2,
                    text="Arquivos obrigatorios"
                    )

label_opc = Label(top_frame2,
                  text="Arquivos opcionais",
                  fg="blue")

# Etiqueta co logo do GTE
photogte = PhotoImage(
    # file=r"C:\Users\Usuario\OneDrive - Universidade de Vigo\Traballo\Machine learning\Probas Biomasa AP\Scripts\imaxes\gte2.png")
    file=r"D:\OneDrive - Universidade de Vigo\Traballo\Machine learning\Probas Biomasa AP\Scripts\imaxes\gte2.png")

photoimagegte = photogte.subsample(3, 3)

label_gte = Label(top_frame2,
                  image=photoimagegte)

# Botóns do Explorador de Arquivos
button_explore1 = Button(top_frame2,
                         text="Selección de arquivo 1",
                         command=browseFiles1)

button_explore2 = Button(top_frame2,
                         text="Selección de arquivo 2",
                         command=browseFiles2)

button_explore3 = Button(top_frame2,
                         text="Selección de arquivo 3",
                         command=browseFiles3,
                         fg="blue")

button_explore4 = Button(top_frame2,
                         text="Selección de arquivo 4",
                         command=browseFiles4,
                         fg="blue")

button_explore5 = Button(top_frame2,
                         text="Selección de arquivo 5",
                         command=browseFiles5,
                         fg="blue")

button_exit = Button(low_frame,
                     text="Saír",
                     fg="red",
                     command=exit)

# Botón de frankenstein
photo = PhotoImage(
    # file=r"C:\Users\Usuario\OneDrive - Universidade de Vigo\Traballo\Machine learning\Probas Biomasa AP\Scripts\imaxes\puzzle.png")
    file=r"D:\OneDrive - Universidade de Vigo\Traballo\Machine learning\Probas Biomasa AP\Scripts\imaxes\puzzle.png")

photoimage = photo.subsample(25, 25)

button_frankenstein = Button(top_frame2,
                             text="Frankenstein",
                             image=photoimage,
                             command=frankenstein)

# Grid method is chosen for placing
# the widgets at respective positions
# in a table like structure by
# specifying rows and columns

label_file_explorer.grid(column=1, row=1, columnspan=3)

button_explore1.grid(column=1, row=2, pady=10, padx=5)
button_explore2.grid(column=2, row=2, pady=10, padx=5)
button_explore3.grid(column=3, row=2, pady=10, padx=5)
button_explore4.grid(column=4, row=2, pady=10, padx=5)
button_explore5.grid(column=5, row=2, pady=10, padx=5)

button_frankenstein.grid(column=6, row=2, pady=10, padx=5)

label_gte.grid(column=7, row=2, padx=5)

label_oblig.grid(column=1, row=3, pady=10, padx=5, columnspan=2)
label_opc.grid(column=3, row=3, pady=10, padx=5, columnspan=3)

top_frame2.grid_columnconfigure(5, weight=4)

button_exit.grid(column=3, row=7, padx=5)

# Let the window wait for any events
window.mainloop()
